import pytest
import logging
import tempfile
import itertools
from pathlib import Path

import pycldf
import openpyxl

from helper_functions import (
    copy_to_temp,
    copy_to_temp_no_bib,
    copy_to_temp_bad_bib,
    empty_copy_of_cldf_wordlist,
)
import lexedata.importer.excel_matrix as f
from lexedata.exporter.cognates import ExcelWriter
from lexedata.importer.cognates import import_cognates_from_excel


@pytest.fixture(
    params=[
        "data/cldf/minimal/cldf-metadata.json",
        "data/cldf/smallmawetiguarani/cldf-metadata.json",
    ]
)
def cldf_wordlist(request):
    return Path(__file__).parent / request.param


@pytest.fixture(
    params=[
        (
            "data/excel/small.xlsx",
            "data/excel/small_cog.xlsx",
            "data/cldf/smallmawetiguarani/cldf-metadata.json",
        ),
        (
            "data/excel/minimal.xlsx",
            "data/excel/minimal_cog.xlsx",
            "data/cldf/minimal/cldf-metadata.json",
        ),
    ]
)
def excel_wordlist(request):
    return (
        Path(__file__).parent / request.param[0],
        Path(__file__).parent / request.param[1],
        empty_copy_of_cldf_wordlist(Path(__file__).parent / request.param[2]),
    )


@pytest.fixture(
    params=[
        copy_to_temp,
        copy_to_temp_no_bib,
        copy_to_temp_bad_bib,
    ]
)
def working_and_nonworking_bibfile(request):
    return request.param


def test_fromexcel_runs(excel_wordlist):
    lexicon, cogsets, (empty_dataset, original) = excel_wordlist
    f.load_dataset(Path(empty_dataset.tablegroup._fname), str(lexicon), str(cogsets))


def test_fromexcel_correct(excel_wordlist):
    lexicon, cogsets, (empty_dataset, original_md) = excel_wordlist
    original = pycldf.Wordlist.from_metadata(original_md)
    # TODO: parameterize original, like the other parameters, over possible
    # test datasets.
    f.load_dataset(Path(empty_dataset.tablegroup._fname), str(lexicon), str(cogsets))
    form_ids_from_excel = {form["ID"] for form in empty_dataset["FormTable"]}
    form_ids_original = {form["ID"] for form in original["FormTable"]}
    cognate_ids_from_excel = {
        cognate["ID"] for cognate in empty_dataset["CognateTable"]
    }
    cognate_ids_original = {form["ID"] for form in original["CognateTable"]}
    assert form_ids_original == form_ids_from_excel, "{:} and {:} don't match.".format(
        empty_dataset["FormTable"]._parent._fname.parent
        / str(empty_dataset["FormTable"].url),
        original["FormTable"]._parent._fname.parent
        / str(empty_dataset["FormTable"].url),
    )
    assert (
        cognate_ids_original == cognate_ids_from_excel
    ), "{:} and {:} don't match.".format(
        empty_dataset["CognateTable"]._parent._fname.parent
        / str(empty_dataset["CognateTable"].url),
        original["CognateTable"]._parent._fname.parent
        / str(empty_dataset["CognateTable"].url),
    )


def test_toexcel_runs(cldf_wordlist, working_and_nonworking_bibfile):
    filled_cldf_wordlist = working_and_nonworking_bibfile(cldf_wordlist)
    writer = ExcelWriter(
        dataset=filled_cldf_wordlist[0],
        database_url=str(filled_cldf_wordlist[1]),
    )
    _, out_filename = tempfile.mkstemp(".xlsx", "cognates")
    writer.create_excel(out_filename)


def test_roundtrip(cldf_wordlist, working_and_nonworking_bibfile):
    filled_cldf_wordlist = working_and_nonworking_bibfile(cldf_wordlist)
    dataset, target = filled_cldf_wordlist
    c_formReference = dataset["CognateTable", "formReference"].name
    c_cogsetReference = dataset["CognateTable", "cognatesetReference"].name
    old_judgements = {
        (row[c_formReference], row[c_cogsetReference])
        for row in dataset["CognateTable"].iterdicts()
    }
    writer = ExcelWriter(dataset)
    _, out_filename = tempfile.mkstemp(".xlsx", "cognates")
    writer.create_excel(out_filename)

    # Reset the existing cognatesets and cognate judgements, to avoid
    # interference with the the data in the Excel file
    dataset["CognateTable"].write([])
    dataset["CognatesetTable"].write([])

    ws = openpyxl.load_workbook(out_filename).active

    import_cognates_from_excel(ws, dataset)

    new_judgements = {
        (row[c_formReference], row[c_cogsetReference])
        for row in dataset["CognateTable"].iterdicts()
    }

    assert new_judgements == old_judgements


def test_roundtrip_separator_column(cldf_wordlist, working_and_nonworking_bibfile):
    """Test whether a CognatesetTable column with separator survives a roundtrip."""
    dataset, target = working_and_nonworking_bibfile(cldf_wordlist)
    dataset.add_columns("CognatesetTable", "CommaSeparatedTags")
    dataset["CognatesetTable", "CommaSeparatedTags"].separator = ","
    c_id = dataset["CognatesetTable", "id"].name

    write_back = list(dataset["CognatesetTable"])
    tags = []
    for tag, row in zip(
        itertools.cycle(
            [["two", "tags"], ["single-tag"], [], ["tag;containing;other;separator"]]
        ),
        write_back,
    ):
        tags.append((row[c_id], tag))
        row["CommaSeparatedTags"] = tag
    dataset.write(CognatesetTable=write_back)

    writer = ExcelWriter(dataset)
    _, out_filename = tempfile.mkstemp(".xlsx", "cognates")
    writer.create_excel(out_filename)

    ws = openpyxl.load_workbook(out_filename).active
    import_cognates_from_excel(ws, dataset)

    reread_tags = [
        (c[c_id], c["CommaSeparatedTags"]) for c in dataset["CognatesetTable"]
    ]
    reread_tags.sort(key=lambda x: x[0])
    tags.sort(key=lambda x: x[0])
    assert reread_tags == tags


def test_cell_comments():
    dataset, _ = copy_to_temp(
        Path(__file__).parent / "data/cldf/minimal/cldf-metadata.json"
    )
    excel_filename = Path(__file__).parent / "data/excel/judgement_cell_with_note.xlsx"

    ws = openpyxl.load_workbook(excel_filename).active
    import_cognates_from_excel(ws, dataset)
    cognates = {
        cog["ID"]: {k: v for k, v in cog.items()} for cog in dataset["CognateTable"]
    }
    assert cognates == {
        "autaa_Woman-cogset": {
            "Cognateset": "cogset",
            "Comment": "Comment on judgement",
            "Form_ID": "autaa_Woman",
            "ID": "autaa_Woman-cogset",
            "Segment_Slice": None,
            "Alignment": None,
        }
    }
    cognatesets = {
        cog["ID"]: {k: v for k, v in cog.items()} for cog in dataset["CognatesetTable"]
    }
    assert cognatesets == {
        "cogset": {"Name": "cogset", "Comment": "Cognateset-comment", "ID": "cogset"}
    }


def test_cell_comments_and_comment_column(caplog):
    dataset, _ = copy_to_temp(
        Path(__file__).parent / "data/cldf/minimal/cldf-metadata.json"
    )
    excel_filename = Path(__file__).parent / "data/excel/judgement_cell_with_note.xlsx"

    sheet = openpyxl.load_workbook(excel_filename).active
    sheet.insert_cols(2)
    sheet.cell(row=1, column=2, value="Comment")
    sheet.cell(row=2, column=2, value="Comment")

    with caplog.at_level(logging.INFO):
        import_cognates_from_excel(sheet, dataset)

    assert "from the cell comments" in caplog.text

    cognates = {
        cog["ID"]: {k: v for k, v in cog.items()} for cog in dataset["CognateTable"]
    }
    assert cognates == {
        "autaa_Woman-cogset": {
            "Cognateset": "cogset",
            "Comment": "Comment on judgement",
            "Form_ID": "autaa_Woman",
            "ID": "autaa_Woman-cogset",
            "Segment_Slice": None,
            "Alignment": None,
        }
    }
    cognatesets = {
        cog["ID"]: {k: v for k, v in cog.items()} for cog in dataset["CognatesetTable"]
    }
    assert cognatesets == {
        "cogset": {"Name": "cogset", "Comment": "Cognateset-comment", "ID": "cogset"}
    }


def test_cell_comments_export():
    dataset, _ = copy_to_temp(
        Path(__file__).parent / "data/cldf/minimal/cldf-metadata.json"
    )
    _, out_filename = tempfile.mkstemp(".xlsx", "cognates")

    E = ExcelWriter(dataset)
    E.set_header()
    E.create_excel(out_filename, size_sort=False, language_order="Name")

    ws_out = openpyxl.load_workbook(out_filename).active
    for col in ws_out.iter_cols():
        pass
    assert col[
        -1
    ].comment.content, "Last row of last column should contain a judgement, with a comment attached to it."
    assert (
        col[-1].comment.content == "A judgement comment"
    ), "Comment should match the comment from the cognate table"
