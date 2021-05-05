"""Import some Bantu data

This script is a very rough sketch for importing yet another shape of data.
Here, every even column contains cells with forms (some cells containing
multiple forms), while the odd columns contain the associated cognate codes
(generally in the same order).

"""
import re
import csv
import sys

import openpyxl

if __name__ == "__main__":
    comma_or_semicolon = re.compile("[,;]\\W*")

    ws = openpyxl.load_workbook("bantu/bantu.xlsx").active

    concepts = []
    for concept_metadata in ws.iter_cols(min_col=1, max_col=1, min_row=2):
        for entry, cogset in zip(concept_metadata[::2], concept_metadata[1::2]):
            try:
                concepts.append(entry.value.strip())
            except AttributeError:
                break

    w = csv.writer(open("forms.csv", "w"))

    w.writerow(["Language_ID", "Concept_ID", "Form", "Comment", "Cognateset"])

    for language in ws.iter_cols(min_col=2):
        language_name = language[0].value
        for c, (entry, cogset) in enumerate(zip(language[1::2], language[2::2])):
            if not entry.value:
                assert not cogset.value
                continue
            bracket_level = 0
            i = 0
            f = entry.value.strip()
            forms = []
            while i < len(f):
                match = comma_or_semicolon.match(f[i:])
                if f[i] == "(":
                    bracket_level += 1
                    i += 1
                    continue
                elif f[i] == ")":
                    bracket_level -= 1
                    i += 1
                    continue
                elif bracket_level:
                    i += 1
                    continue
                elif match:
                    forms.append(f[:i].strip())
                    i += match.span()[1]
                    f = f[i:]
                    i = 0
                else:
                    i += 1

            forms.append(f.strip())

            if type(cogset.value) == float:
                cogsets = [str(int(cogset.value))]
            else:
                cogsets = comma_or_semicolon.split(cogset.value.strip())

            if len(cogsets) == 1 or len(cogsets) == len(forms):
                True
            else:
                print(
                    "{:}: Forms ({:}) did not match cognates ({:})".format(
                        entry.coordinate, ", ".join(forms), ", ".join(cogsets)
                    ),
                    file=sys.stderr,
                )

            for form, cogset in zip(forms, cogsets + [None]):
                w.writerow([language_name, concepts[c], form, None, cogset])
